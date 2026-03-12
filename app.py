import sqlite3
import requests
import re
import json # Added for parsing JSON from LLM
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, redirect, session, Response, jsonify
from openpyxl import Workbook
from groq import Groq # Added Groq import
import os
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
from ai_agent import generate_newsletter_draft
from cross_project_matcher import find_matching_projects
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "leads.db")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "project_logos")

load_dotenv(os.path.join(BASE_DIR, ".env"))  # Always load env from project folder
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__)
app.secret_key = "tars_stable_system"
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
SERP_API_KEY = os.getenv("SERP_API_KEY")

@app.context_processor
def inject_active_page():
    from flask import request
    path = request.path.lower()

    if path.startswith("/dashboard"):
        active = "dashboard"

    elif path.startswith("/chat") or path.startswith("/chat_session"):
        # 👈 THIS LINE FIXES AI Chat highlight
        active = "chat"

    elif path.startswith("/overview"):
        active = "overview"

    elif path.startswith("/savedprojects"):
        active = "product"

    elif path.startswith("/call_for_action"):
        active = "cta"

    elif path.startswith("/industry_viewed"):
        active = "industry"

    else:
        active = ""

    return dict(active_page=active)

# Initialize Groq Client

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
client = Groq(api_key=GROQ_API_KEY)
def get_db_connection():
    return sqlite3.connect(DB_PATH)

# ===============================
# USER MODEL (Flask-Login)
# ===============================

class User(UserMixin):

    def __init__(self, id, username):
        self.id = id
        self.username = username
@login_manager.user_loader
def load_user(user_id):

    conn = get_db_connection()

    user = conn.execute(
        "SELECT id, username FROM users WHERE id=?",
        (user_id,)
    ).fetchone()

    conn.close()

    if user:
        return User(user[0], user[1])

    return None
# ===============================
# DATABASE INIT (COMPANIES TABLE)
# ===============================
def init_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        website TEXT,
        phone TEXT,
        address TEXT,
        rating TEXT,
        chat_id INTEGER
    )
    """)

    # Safely add new columns if not exist
    for col in [
        "description", "email", "ceo",
        "company_linkedin", "leadership_linkedin", "status"
    ]:
        try:
            cur.execute(f"ALTER TABLE companies ADD COLUMN {col} TEXT")
        except:
            pass
    # ensure chat_id exists
    try:
        cur.execute("ALTER TABLE companies ADD COLUMN chat_id INTEGER")
    except:
        pass

    # Safely add chat_id column
    try:
        cur.execute("ALTER TABLE companies ADD COLUMN chat_id INTEGER")
    except:
        pass

    

    cur.execute("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE,
    password TEXT
)
""")
    conn.commit()
    conn.close()

# ===============================
# CHAT SYSTEM TABLES (NEW)
# ===============================
def init_chat_system():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS chats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user TEXT,
        title TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chat_id INTEGER,
        role TEXT,
        content TEXT,
        timestamp DEFAULT CURRENT_TIMESTAMP
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chat_id INTEGER,
        product_name TEXT,
        description TEXT,
        industry TEXT,
        location TEXT,
        industry_suggestions TEXT
    )
    """)

    # add column safely if missing
    try:
        cur.execute("ALTER TABLE products ADD COLUMN industry_suggestions TEXT")
    except:
        pass

    # cleanup potential test project
    try:
        cur.execute("DELETE FROM chats WHERE title = 'Test'")
        cur.execute("DELETE FROM products WHERE chat_id NOT IN (SELECT id FROM chats)")
    except:
        pass

    conn.commit()
    conn.close()


# Run both initializations
init_db()
init_chat_system()

# ===============================
# WEBSITE SCRAPER
# ===============================
def extract_website_data(url):
    try:
        html = requests.get(url, timeout=5).text
        soup = BeautifulSoup(html, "html.parser")

        emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]+", html)
        email = emails[0] if emails else ""

        desc = ""
        meta = soup.find("meta", attrs={"name": "description"})
        if meta:
            desc = meta.get("content", "")
        elif soup.title:
            desc = soup.title.text.strip()

        return email, desc
    except:
        return "", ""

# ===============================
# CEO LINKEDIN
# ===============================
def find_ceo_with_linkedin(company_name):
    try:
        query = f"{company_name} CEO LinkedIn"

        params = {"q": query, "engine": "google", "api_key": SERP_API_KEY}
        data = requests.get("https://serpapi.com/search", params=params).json()

        for result in data.get("organic_results", []):
            link = result.get("link", "")
            title = result.get("title", "")

            if "linkedin.com/in/" in link:
                name = title.split("-")[0].strip()
                return name, link

        return "Not Available", ""
    except:
        return "Not Available", ""

# ===============================
# COMPANY + HR LINKEDIN (NEW)
# ===============================
def find_company_and_hr_linkedin(company_name):
    try:
        query = f"{company_name} company LinkedIn"
        params = {"q": query, "engine": "google", "api_key": SERP_API_KEY}
        data = requests.get("https://serpapi.com/search", params=params).json()

        company_linkedin = ""
        leadership_linkedin = ""

        for result in data.get("organic_results", []):
            link = result.get("link", "")

            if "linkedin.com/company/" in link and not company_linkedin:
                company_linkedin = link

            if "linkedin.com/in/" in link and not leadership_linkedin:
                leadership_linkedin = link

        return company_linkedin, leadership_linkedin
    except:
        return "", ""

# ===============================
# GOOGLE MAPS SEARCH
# ===============================
def map_industry_to_search(industry):
    return f"{industry}"

def search_companies_maps(keyword, city):
    params = {
        "engine": "google_maps",
        "q": f"{keyword} company in {city}",
        "api_key": SERP_API_KEY
    }

    data = requests.get("https://serpapi.com/search", params=params).json()
    companies = []

    for r in data.get("local_results", []):
        name = r.get("title")
        website = r.get("website")
        phone = r.get("phone", "")
        address = r.get("address", "")
        rating = r.get("rating", "")
        description = r.get("description", "")

        email = ""
        ceo = "Not Available"

        if website:
            if not website.startswith("http"):
                website = "https://" + website
            email, desc2 = extract_website_data(website)
            if not description:
                description = desc2

        # CEO enrichment
        ceo_name, ceo_link = find_ceo_with_linkedin(name)
        if ceo_link:
            ceo = f"{ceo_name}|{ceo_link}"

        # NEW LinkedIn enrichment
        company_ln, hr_ln = find_company_and_hr_linkedin(name)

        companies.append({
            "name": name,
            "website": website,
            "phone": phone,
            "address": address,
            "rating": rating,
            "description": description,
            "email": email,
            "ceo": ceo,
            "company_linkedin": company_ln,
            "leadership_linkedin": hr_ln
        })

    return companies

# ===============================
# DB OPS
# ===============================
def clear_companies(chat_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM companies WHERE chat_id = ?", (chat_id,))
    conn.commit()
    conn.close()

def save_company(c, chat_id):

    conn = get_db_connection()
    conn.execute("""
INSERT INTO companies
(name, website, phone, address, rating, description,
 email, ceo, company_linkedin, leadership_linkedin, chat_id)
VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
""", (
    c["name"],
    c["website"],
    c["phone"],
    c["address"],
    c["rating"],
    c["description"],
    c.get("email"),
    c.get("ceo"),
    c.get("company_linkedin"),
    c.get("leadership_linkedin"),
    chat_id      # ⭐ THIS IS THE IMPORTANT PART
))
    conn.commit()
    conn.close()


def get_companies(chat_id):
    conn = get_db_connection()

    rows = conn.execute("""
    SELECT name, website, phone, address, rating, description,
           email, ceo, company_linkedin, leadership_linkedin
    FROM companies
    WHERE chat_id = ?
    """, (chat_id,)).fetchall()

    conn.close()

    return [{
        "name": r[0],
        "website": r[1],
        "phone": r[2],
        "address": r[3],
        "rating": r[4],
        "description": r[5],
        "email": r[6],
        "ceo": r[7],
        "company_linkedin": r[8],
        "leadership_linkedin": r[9]
    } for r in rows]


# ===============================
# SEARCH EXISTING COMPANIES IN DB
# ===============================
def search_existing_companies(city):
    """
    Fetch companies already stored in DB that match the selected city.
    We exclude duplicates by company name.
    """
    conn = get_db_connection()

    rows = conn.execute("""
        SELECT DISTINCT name, website, phone, address, rating,
               description, email, ceo, company_linkedin, leadership_linkedin
        FROM companies
        WHERE address LIKE ?
    """, (f"%{city}%",)).fetchall()

    conn.close()

    companies = [{
        "name": r[0],
        "website": r[1],
        "phone": r[2],
        "address": r[3],
        "rating": r[4],
        "description": r[5],
        "email": r[6],
        "ceo": r[7],
        "company_linkedin": r[8],
        "leadership_linkedin": r[9]
    } for r in rows]

    return companies

# ===============================
# AI RELEVANCE FILTER
# ===============================
def filter_relevant_companies(project_description, companies):
    """
    Uses Groq LLM to check which companies are relevant to the project.
    Returns only relevant companies.
    """
    try:

        context = "\n".join([
            f"{c['name']} - {c.get('description','')}"
            for c in companies[:30]  # limit for token safety
        ])

        prompt = f"""
        Project Description:
        {project_description}

        Companies:
        {context}

        From the companies list, return ONLY the company names that are relevant
        to the project.

        Return ONLY a JSON list.

        Example:
        ["Apollo Hospital", "HealthPlix"]
        """

        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            max_tokens=200
        )

        response = completion.choices[0].message.content.strip()

        start = response.find("[")
        end = response.rfind("]") + 1

        if start == -1:
            return []

        names = json.loads(response[start:end])

        relevant = [c for c in companies if c["name"] in names]

        return relevant

    except Exception as e:
        print("Relevance filter error:", e)
        return []
# ===============================
# INDUSTRY SUGGESTION ENGINE
# ===============================
def suggest_industries(description):
    """
    Uses Llama 3.1 8B via Groq to dynamically detect industries 
    based on the user's product description.
    """
    try:
        

        prompt = f"""
        Analyze the following product or service description and identify the most relevant target industries.

        Description:
        "{description}"

        Return ONLY a Python-style list of 3-6 industry names that would be ideal target markets.

        Example format:
        ["Healthcare", "FinTech", "Logistics"]

        Do not explain. Do not add commentary. Only return the list.
        """

        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that categorizes products into industries for lead generation."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1, # Low temperature for consistent categorization
            max_tokens=100
        )

        # Extract content and attempt to parse the list
        response_text = completion.choices[0].message.content.strip()
        
        # Simple cleanup to handle potential LLM conversational fluff
        start_idx = response_text.find('[')
        end_idx = response_text.find(']') + 1
        if start_idx != -1 and end_idx != -1:
            industries = json.loads(response_text[start_idx:end_idx].replace("'", '"'))
            return industries
        
        return ["General Business"]
    except Exception as e:
        print(f"Groq API Error: {e}")
        return ["General Business"]


# ---------------------
# HTML formatting helper
# ---------------------

def format_chat_html(text):
    """Escape and convert simple plaintext into HTML suitable for chat bubbles.
    Preserves line breaks and converts bullets (-,*,🔹,✅) into list items with an icon.
    """
    import html
    escaped = html.escape(text)
    lines = escaped.splitlines()
    out = []
    in_ul = False
    for line in lines:
        m = re.match(r'^\s*[-*]\s+(.*)', line)
        e = re.match(r'^\s*[🔹✅•]\s*(.*)', line)
        if m or e:
            if not in_ul:
                out.append('<ul>')
                in_ul = True
            item = (m or e).group(1)
            out.append(f'<li><i class="fas fa-circle" style="font-size:0.6em;margin-right:4px;"></i>{item}</li>')
        else:
            if in_ul:
                out.append('</ul>')
                in_ul = False
            out.append(line)
    if in_ul:
        out.append('</ul>')
    return '<br>'.join(out)

def chat_html_to_text(content):
    """Convert stored chat HTML into plain text before sending to the model."""
    if not content:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", content, flags=re.I)
    text = re.sub(r"</?(ul|ol)\b[^>]*>", "\n", text, flags=re.I)
    text = re.sub(r"<li\b[^>]*>", "- ", text, flags=re.I)
    text = re.sub(r"</li>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    import html as _html
    text = _html.unescape(text)
    return re.sub(r"\n{3,}", "\n\n", text).strip()

# ===============================
# CHAT HELPERS (NEW)
# ===============================
def generate_sales_ai_reply(chat_id, user_message):
    try:
        previous_messages = get_chat_messages(chat_id)
        product_name, product_description, _ = get_product_info(chat_id)

        conversation = [
            {
                "role": "system",
                "content": (
                    "You are a highly interactive AI sales strategist and startup advisor. "
                    "Understand the user's message semantically even when phrased in different words. "
                    "Use conversation context to give specific, non-repetitive answers. "
                    "Never act like a rigid form bot and never ask for magic trigger phrases. "
                    "If details are missing, ask one short follow-up question. "
                    "Keep responses practical, clear, and conversational."
                ),
            }
        ]

        if product_name or product_description:
            product_context = (
                f"Known project context:\n"
                f"- Product name: {product_name or 'Not provided'}\n"
                f"- Product description: {product_description or 'Not provided'}"
            )
            conversation.append({"role": "system", "content": product_context})

        for role, content, _ in previous_messages[-20:]:
            if role not in ("user", "assistant", "system"):
                continue
            clean_content = chat_html_to_text(content)
            if clean_content:
                conversation.append({"role": role, "content": clean_content})

        conversation.append({"role": "user", "content": user_message.strip()})

        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=conversation,
            temperature=0.5,
            max_tokens=500,
        )

        return completion.choices[0].message.content.strip()
    except Exception as e:
        print("Sales AI Error:", e)
        return "Tell me a bit more about what you're building and what kind of help you want."


def get_product_info(chat_id):
    """Return (product_name, description, industry_suggestions_json) for a chat if present."""
    conn = get_db_connection()
    row = conn.execute(
        "SELECT product_name, description, industry_suggestions FROM products WHERE chat_id=?",
        (chat_id,)
    ).fetchone()
    conn.close()
    if row:
        return row[0], row[1], row[2]
    return None, None, None


def upsert_product(chat_id, product_name=None, description=None, industry_suggestions=None):
    """Update the products row for chat_id if exists, otherwise insert one.
    Only non-None fields are updated.  industry_suggestions should be JSON text.
    """
    conn = get_db_connection()
    cur = conn.cursor()

    existing = cur.execute(
        "SELECT id FROM products WHERE chat_id=?",
        (chat_id,)
    ).fetchone()

    if existing:
        updates = []
        params = []
        if product_name is not None:
            updates.append("product_name=?")
            params.append(product_name)
        if description is not None:
            updates.append("description=?")
            params.append(description)
        if industry_suggestions is not None:
            updates.append("industry_suggestions=?")
            params.append(industry_suggestions)
        if updates:
            params.append(chat_id)
            cur.execute(f"UPDATE products SET {', '.join(updates)} WHERE chat_id=?", params)
    else:
        cur.execute(
            "INSERT INTO products (chat_id, product_name, description, industry_suggestions) VALUES (?, ?, ?, ?)",
            (chat_id, product_name or "", description or "", industry_suggestions or "")
        )

    conn.commit()
    conn.close()


def user_mentions_project(msg):
    s = msg.lower()
    keywords = ["building", "build", "built", "working on", "develop", "developing", "project", "launch"]
    return any(k in s for k in keywords)


def extract_project_name(msg):
    """Try to pull a project name from an informal sentence.
    Uses simple regex heuristics; returns None if unsure.
    """
    s = msg.strip()
    # look for explicit phrases
    m = re.search(r"(?:called|named)\s+([A-Za-z0-9 _-]+)", s, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b(?:building|built|developing|working on)\s+(?:a |an |the )?([A-Za-z0-9 _-]+)", s, re.I)
    if m:
        return m.group(1).strip()
    return None


def user_requests_industry_suggestions(msg):
    s = msg.lower()
    # if the message mentions industry/market and also contains words that
    # imply a request (suggest, idea, recommend, target, where, which)
    if not ("industry" in s or "market" in s):
        return False
    keywords = ["suggest", "idea", "recommend", "target", "where", "which", "ideas"]
    return any(k in s for k in keywords)


def extract_industries_from_text(text):
    """
    Extract industries ONLY from properly formatted list items
    (numbered or bullet style). Do NOT extract capitalized phrases
    to avoid capturing project names.
    """
    industries = []

    for line in text.splitlines():
        line = line.strip()
        clean = line.replace("**", "").replace("*", "")

        # Match numbered or bulleted items ending with colon
        m = re.match(r'^(?:\d+\.|[-*•])\s*([^:]+):', clean)
        if m:
            industries.append(m.group(1).strip())

    # Deduplicate while preserving order
    seen = set()
    result = []
    for ind in industries:
        if ind not in seen:
            seen.add(ind)
            result.append(ind)

    return result

def create_chat(user, title):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO chats (user, title) VALUES (?, ?)", (user, title))
    chat_id = cur.lastrowid
    conn.commit()
    conn.close()
    return chat_id


def save_message(chat_id, role, content):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO messages (chat_id, role, content) VALUES (?, ?, ?)",
        (chat_id, role, content)
    )
    conn.commit()
    conn.close()


def get_user_chats():

    conn = get_db_connection()

    rows = conn.execute(
        "SELECT id, title FROM chats WHERE user=? ORDER BY id DESC",
         (current_user.username,)
    ).fetchall()

    conn.close()

    return rows


def get_chat_messages(chat_id):

    conn = get_db_connection()

    rows = conn.execute("""
        SELECT m.role, m.content, m.timestamp
        FROM messages m
        JOIN chats ch ON m.chat_id = ch.id
        WHERE m.chat_id = ? AND ch.user = ?
        ORDER BY m.id ASC
    """, (chat_id, current_user.username)).fetchall()

    conn.close()

    return rows


# ===============================
# CROSS-PROJECT INTELLIGENCE (NEW)
# ===============================

@app.route("/industry_viewed")
@login_required
def industry_viewed():
   
    
    conn = get_db_connection()
    # Fetch projects and their associated companies
    # We join chats -> products -> companies
    data = conn.execute("""
        SELECT 
            p.product_name, 
            p.description, 
            c.name, 
            c.description, 
            c.rating
        FROM products p
        JOIN chats ch ON p.chat_id = ch.id
        LEFT JOIN companies c ON ch.id = c.chat_id
        WHERE ch.user = ?
    """,(current_user.username,)).fetchall()
    conn.close()

    if not data:
        return render_template(
    "industry_viewed.html",
    analysis="No leads or projects found to analyze yet.",
    active_page="industry"
)

    # Organize data for the LLM
    structured_context = {}
    for p_name, p_desc, comp_name, comp_desc, rating in data:
        if p_name not in structured_context:
            structured_context[p_name] = {"desc": p_desc, "leads": []}
        if comp_name:
            structured_context[p_name]["leads"].append(f"{comp_name} (Rating: {rating}) - {comp_desc}")

    # Build the text prompt
    context_text = "Here is the user's sales portfolio and the specific companies they have scraped:\n"
    for project, info in structured_context.items():
        context_text += f"\nPROJECT: {project}\nOFFERING: {info['desc']}\nTARGETED COMPANIES:\n"
        context_text += "\n".join(info['leads'][:10]) # Limit to top 10 leads per project for token space
        context_text += "\n"

    prompt = f"""
    Analyze the following sales projects and companies:

    {context_text}

    For each company:
    - Identify the BEST matching project(s)
    - Return ONLY valid JSON in this format:

    [
    {{
        "company": "Company Name",
        "projects": ["Project 1", "Project 2"]
    }}
    ]

    Return ONLY JSON. No explanation. No HTML.
    """
    
    try:
        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7
        )
        response_text = completion.choices[0].message.content.strip()

        start = response_text.find("[")
        end = response_text.rfind("]") + 1

        if start != -1 and end != -1:
            json_text = response_text[start:end]
            analysis_data = json.loads(json_text)
        else:
            analysis_data = []
    except Exception as e:
        analysis_data = []
        print(f"Industry analysis error: {e}")

    return render_template("industry_viewed.html", analysis=analysis_data, active_page="industry")


# ===============================
# ROUTES (UNCHANGED)
# ===============================
@app.route("/signup", methods=["GET","POST"])
def signup():

    if request.method == "POST":

        name = request.form["name"]
        phone = request.form["phone"]
        email = request.form["email"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        company_name = request.form["company_name"]
        employees = request.form["employees"]
        location = request.form["location"]
        company_email = request.form["company_email"]
        company_phone = request.form["company_phone"]

        # password check
        if password != confirm_password:
            return "Passwords do not match"

        hashed_password = generate_password_hash(password)

        conn = get_db_connection()

        try:

            conn.execute("""
            INSERT INTO users (username,password)
            VALUES (?,?)
            """,(email,hashed_password))

            conn.commit()

        except:
            conn.close()
            return "User already exists"

        user = conn.execute(
            "SELECT id, username FROM users WHERE username=?",
            (email,)
        ).fetchone()

        conn.close()

        # auto login
        user_obj = User(user[0], user[1])
        login_user(user_obj)

        return redirect("/dashboard")

    return render_template("signup.html")


@app.route("/", methods=["GET","POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()

        user = conn.execute(
            "SELECT * FROM users WHERE username=?",
            (username,)
        ).fetchone()

        conn.close()

        if not user:
            return "You don't have an account. Please sign up first."

        if not check_password_hash(user[2], password):
            return "Incorrect password."

        user_obj = User(user[0], user[1])
        login_user(user_obj)

        conn = get_db_connection()

        row = conn.execute("""
        SELECT id
        FROM chats
        WHERE user = ?
        ORDER BY id DESC
        LIMIT 1
        """, (username,)).fetchone()

        conn.close()

        if row:
            session["active_chat"] = row[0]
        else:
            session.pop("active_chat", None)

        session.pop("suggested_industry", None)
        session.pop("selected_city", None)

        return redirect("/dashboard")

    return render_template("login.html")
# ===============================
# CHAT ROUTES (NEW)
# ===============================

@app.route("/chat")
@login_required
def chat_home():

    chats = get_user_chats()

    active_chat = session.get("active_chat")

    return render_template(
        "chat.html",
        chats=chats,
        show_create=False,
        active_chat=active_chat,
        active_page="chat"
    )

@app.route("/create_project")
@login_required
def create_project_page():
    

    chats = get_user_chats()

    return render_template(
        "chat.html",
        chats=chats,
        show_create=True,
        active_page="chat"
    )

@app.route("/new_chat", methods=["POST"])
@login_required
def new_chat():
   

    title = request.form.get("title", "New Project")
    description = request.form.get("description", "")

    # 1️⃣ Create chat first
    chat_id = create_chat(current_user.username, title)

    # 2️⃣ Save product entry
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO products (chat_id, product_name, description) VALUES (?, ?, ?)",
        (chat_id, title, description)
    )
    conn.commit()
    conn.close()

    # 3️⃣ Save uploaded logo
    logo = request.files.get("logo")
    if logo and logo.filename != "":
        # Backend validation: allow only PNG/JPEG by MIME type and extension
        filename_raw = secure_filename(logo.filename)
        ext = filename_raw.rsplit('.', 1)[-1].lower() if '.' in filename_raw else ''
        allowed_ext = {'png', 'jpg', 'jpeg'}
        allowed_mimes = {'image/png', 'image/jpeg'}
        mime = getattr(logo, 'mimetype', '') or getattr(logo, 'content_type', '')

        if ext not in allowed_ext or mime not in allowed_mimes:
            # rollback product/chat created earlier to avoid orphan rows
            try:
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("DELETE FROM products WHERE chat_id=?", (chat_id,))
                cur.execute("DELETE FROM chats WHERE id=?", (chat_id,))
                conn.commit()
                conn.close()
            except Exception:
                pass
            return "Invalid file type. Only PNG and JPG images are allowed.", 400

        # save file with chat_id-based filename preserving extension
        filename = secure_filename(f"{chat_id}.{ext}")
        logo_path = os.path.join(UPLOAD_FOLDER, filename)
        logo.save(logo_path)

    return redirect(f"/chat_session/{chat_id}")



@app.route("/chat_session/<int:chat_id>")
@login_required
def chat_session(chat_id):
   

    messages = get_chat_messages(chat_id)
    chats = get_user_chats()
    session["active_chat"] = chat_id

    return render_template(
        "chat.html",
        messages=messages,
        chats=chats,
        active_chat=chat_id,
        show_create=False,
        active_page="chat"
    )


@app.route("/send_message", methods=["POST"])
def send_message():
    chat_id = int(request.form["chat_id"])
    message = request.form.get("message", "").strip()
    if not message:
        return redirect(f"/chat_session/{chat_id}")

    conn = get_db_connection()
    row = conn.execute(
        "SELECT 1 FROM chats WHERE id=? AND user=?",
        (chat_id, current_user.username)
    ).fetchone()
    conn.close()
    if not row:
        return redirect("/chat")

    save_message(chat_id, "user", format_chat_html(message))

    ai_reply = generate_sales_ai_reply(chat_id, message)
    save_message(chat_id, "assistant", format_chat_html(ai_reply))

    inds = extract_industries_from_text(ai_reply)
    if inds:
        session["suggested_industry"] = inds
        import json as _json
        upsert_product(chat_id, industry_suggestions=_json.dumps(inds))

    return redirect(f"/chat_session/{chat_id}")


@app.route("/delete_project/<int:chat_id>", methods=["POST"])
@login_required
def delete_project(chat_id):
   

    conn = get_db_connection()
    cur = conn.cursor()

    # Delete related data first (to avoid orphan records)
    cur.execute("DELETE FROM messages WHERE chat_id = ?", (chat_id,))
    cur.execute("DELETE FROM products WHERE chat_id = ?", (chat_id,))
    cur.execute("DELETE FROM companies WHERE chat_id = ?", (chat_id,))
    cur.execute("DELETE FROM chats WHERE id = ?", (chat_id,))

    conn.commit()
    conn.close()

    # Delete logo file if exists
    logo_path = os.path.join(UPLOAD_FOLDER, f"{chat_id}.png")
    if os.path.exists(logo_path):
        os.remove(logo_path)

    return "", 204

   
# ===============================
# RUN TARGETING (FROM DASHBOARD BUTTON)
# ===============================
@app.route("/run_targeting", methods=["POST"])
@login_required
def run_targeting():
    

    selected_industries = request.form.getlist("industry[]")
    city = request.form["city"]

    chat_id = session.get("active_chat")

    # clear previous project leads
    clear_companies(chat_id)

    # get project description
    product_name, project_description, _ = get_product_info(chat_id)

    # ----------------------------
    # STEP 1: SEARCH EXISTING DB
    # ----------------------------
    db_companies = search_existing_companies(city)

    print("DB companies found:", len(db_companies))

    # ----------------------------
    # STEP 2: AI FILTER
    # ----------------------------
    relevant_companies = filter_relevant_companies(
        project_description or "",
        db_companies
    )

    print("Relevant DB companies:", len(relevant_companies))

    for c in relevant_companies:
        save_company(c, chat_id)

    # ----------------------------
    # STEP 3: SERP SEARCH IF NEEDED
    # ----------------------------
    lead_threshold = 30

    current_leads = len(relevant_companies)

    if current_leads < lead_threshold:

        for ind in selected_industries:

            search_keyword = map_industry_to_search(ind)

            companies = search_companies_maps(search_keyword, city)

            for c in companies:
                save_company(c, chat_id)

    session["selected_city"] = city
    session["suggested_industry"] = selected_industries

    return redirect("/dashboard")





# ===============================
# DASHBOARD PAGE
# ===============================
@app.route("/dashboard")
@login_required
def dashboard():
    

    chat_id = session.get("active_chat")
    companies = get_companies(chat_id) if chat_id else []

    # load projects list
    conn = get_db_connection()
    projects = conn.execute("""
        SELECT ch.id, ch.title, p.description
        FROM chats ch
        LEFT JOIN products p ON ch.id = p.chat_id
        WHERE ch.user = ?
        ORDER BY ch.id DESC
    """, (current_user.username,)).fetchall()
    conn.close()

    suggested = []
    if chat_id:
        _, _, stored = get_product_info(chat_id)
        if stored:
            try:
                suggested = json.loads(stored)
            except:
                suggested = []

    

    city = session.get("selected_city", "")

    return render_template(
        "dashboard.html",
        companies=companies,
        suggested_industry=suggested,
        city=city,
        projects=projects,   # ⭐ NEW
        active_page="dashboard"
    )


@app.route("/savedprojects")
@login_required
def saved_projects():
    

    conn = get_db_connection()
    rows = conn.execute("""
        SELECT ch.id, ch.title, p.description
        FROM chats ch
        LEFT JOIN products p ON ch.id = p.chat_id
        WHERE ch.user = ?
        ORDER BY ch.id DESC
    """, (current_user.username,)).fetchall()
    conn.close()

    return render_template("savedprojects.html", chats=rows, active_page="product")

# ===============================
# OPEN SELECTED PROJECT
# ===============================
@app.route("/open_project/<int:chat_id>")
@login_required
def open_project(chat_id):
    

    # store selected project in session
    session["active_chat"] = chat_id

    return redirect("/dashboard")

@app.route("/export_excel")
@login_required
def export_excel():
   

    chat_id = session.get("active_chat")
    if not chat_id:
        return redirect("/dashboard")

    companies = get_companies(chat_id)

    conn = get_db_connection()
    row = conn.execute(
        "SELECT title FROM chats WHERE id=? AND user=?",
        (chat_id, current_user.username)
    ).fetchone()
    conn.close()

    if not row:
        return redirect("/dashboard")

    title = row[0]

    file_path = "company_leads.xlsx"

    if os.path.exists(file_path):
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

    # Create new sheet with project name
    ws = wb.create_sheet(title=title[:30])  # Excel max 31 chars

    ws.append([
        "Company","Website","Phone","Address","Rating",
        "Description","Email","CEO","Company LinkedIn","Leadership LinkedIn"
    ])

    for c in companies:
        ws.append(list(c.values()))

    wb.save(file_path)

    return Response(open(file_path, "rb"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=company_leads.xlsx"})

@app.route("/call_for_action")
@login_required
def call_for_action():
    

    conn = get_db_connection()
    rows = conn.execute("""
        SELECT ch.id, ch.title, p.description
        FROM chats ch
        LEFT JOIN products p ON ch.id = p.chat_id
        WHERE ch.user = ?
        ORDER BY ch.id DESC
    """,(current_user.username,)).fetchall()

    projects = []
    for chat_id, title, description in rows:
        company_rows = conn.execute("""
            SELECT name, description, rating, email, phone, address, website
            FROM companies
            WHERE chat_id = ? AND status = 'cta'
            ORDER BY id DESC
        """, (chat_id,)).fetchall()

        if company_rows:
            projects.append({
                "title": title,
                "description": description,
                "companies": [{
                    "name": c[0],
                    "description": c[1],
                    "rating": c[2],
                    "email": c[3],
                    "phone": c[4],
                    "address": c[5],
                    "website": c[6]
                } for c in company_rows]
            })

    conn.close()

    return render_template(
    "call_for_action.html",
    projects=projects,
    active_page="cta"
)
@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect("/")

@app.route("/overview")
@login_required
def overview():
   

    conn = get_db_connection()
    rows = conn.execute("""
        SELECT ch.id, ch.title, p.description
        FROM chats ch
        LEFT JOIN products p ON ch.id = p.chat_id
        WHERE ch.user = ?
        ORDER BY ch.id DESC
    """,(current_user.username,)).fetchall()
    conn.close()

    return render_template("overview.html", projects=rows, active_page="overview")


@app.route("/overview_project/<int:chat_id>")
@login_required
def overview_project(chat_id):
   

    conn = get_db_connection()

    companies = conn.execute("""
    SELECT c.id, c.name, c.website, c.phone, c.address, c.rating,
           c.description, c.email, c.ceo,
           c.company_linkedin, c.leadership_linkedin, c.status
    FROM companies c
    JOIN chats ch ON c.chat_id = ch.id
    WHERE c.chat_id = ? AND ch.user = ?
""", (chat_id, current_user.username)).fetchall()

    conn.close()

    # load product info so we can show suggested industries in the view
    pname, pdesc, stored = get_product_info(chat_id)
    suggested = []
    if stored:
        try:
            suggested = json.loads(stored)
        except:
            suggested = []

    return render_template(
        "overview_project.html",
        companies=companies,
        chat_id=chat_id,
        product_name=pname,
        product_description=pdesc,
        suggested_industry=suggested,
        active_page="overview"
    )

@app.route("/update_status", methods=["POST"])
def update_status():
    company_id = request.form["company_id"]
    status = request.form["status"]

    conn = get_db_connection()
    conn.execute(
        "UPDATE companies SET status=? WHERE id=?",
        (status, company_id)
    )
    conn.commit()
    conn.close()

    return "", 204

@app.route("/generate_newsletter", methods=["POST"])
@login_required
def generate_newsletter():
   

    project_title = request.form["project_title"]
    project_description = request.form["project_description"]
    company_name = request.form["company_name"]
    company_description = request.form.get("company_description", "")

    result = generate_newsletter_draft(
        project_title,
        project_description,
        company_name,
        company_description
    )

    return jsonify(result)

@app.route("/send_newsletter", methods=["POST"])
@login_required
def send_newsletter():
    
    
    try:
        data = request.get_json()
        recipient_email = data.get("recipient_email")
        subject = data.get("subject")
        body = data.get("body")
        company_name = data.get("company_name", "")
        
        if not recipient_email or not subject or not body:
            return jsonify({"success": False, "message": "Missing required fields"}), 400
        
        # Get SMTP configuration from environment variables
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        sender_email = os.getenv("SENDER_EMAIL")
        sender_password = os.getenv("SENDER_PASSWORD")
        
        if not sender_email or not sender_password:
            return jsonify({
                "success": False, 
                "message": "Email configuration not set. Please configure SMTP settings in .env file."
            }), 500
        
        # Create message
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = sender_email
        msg["To"] = recipient_email
        
        # Create plain text and HTML versions
        text_part = MIMEText(body, "plain")
        msg.attach(text_part)
        
        # Send email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        
        return jsonify({
            "success": True, 
            "message": f"Newsletter sent successfully to {company_name} ({recipient_email})!"
        })
        
    except smtplib.SMTPAuthenticationError:
        return jsonify({
            "success": False, 
            "message": "Email authentication failed. Please check your email credentials."
        }), 500
    except Exception as e:
        return jsonify({
            "success": False, 
            "message": f"Failed to send email: {str(e)}"
        }), 500

@app.route("/get_company_project_matches")
@login_required
def get_company_project_matches():
   

    company_id = request.args.get("company_id")

    conn = get_db_connection()

    company = conn.execute("""
        SELECT name, description
        FROM companies
        WHERE id = ?
    """, (company_id,)).fetchone()

    if not company:
        conn.close()
        return jsonify([])

    company_name, company_description = company

    # Get all projects of this user
    projects = conn.execute("""
        SELECT ch.title, p.description
        FROM chats ch
        LEFT JOIN products p ON ch.id = p.chat_id
        WHERE ch.user = ?

        ORDER BY ch.id DESC
""", (current_user.username,)).fetchall()

    conn.close()

    matches = find_matching_projects(
        company_name,
        company_description or "",
        projects
    )

    return jsonify(matches)

if __name__ == "__main__":
    app.run(debug=True)
